# -*- coding: utf-8 -*-
## @file xls_net.py
## @brief mark net for net value in a column
## @author jiayanming
import pdb
import os.path
import re
import math
from openpyxl import *
#from openpyxl import styles
from bisect import bisect_left

xls_input = "d:/data/xls_f/216336FY19.XLSX"
xls_output = "d:/data/xls_f/216336FY19.XLSX_output.XLSX"


my_color = styles.colors.Color(rgb="ffff00")
around_color = styles.fills.PatternFill(patternType='solid', fgColor=my_color)

# global variables
idx_column = 11


# ##################### FILE LOAD ####################    
val_list = [0]

def load_value(input_table):
    ws = input_table.active
    max_ver_col = idx_column + 1
    rows = ws.iter_rows(min_row=2, max_col=max_ver_col, values_only=True)
    rid = 0
    for r in rows:
        rid += 1
        if rid == 1:
            print("first value: {}".format(r[idx_column]))            
        val = float(r[idx_column])
        find_net = False
        for i in range(1, len(val_list)):
            if abs(val_list[i] + val) < 1e-3:
                # find net
                val_list[i] = 0
                find_net = True
                break
        if find_net:
            val_list.append(0)
        else:
            val_list.append(val)
        
# write results to ap table 
def update_table():
    ws = table_input.active
    ws.insert_cols(1)
    for ri in range(1, len(val_list)):
        if abs(val_list[ri]) < 1e-3:
            ws.cell(row=ri+1, column=1).value = "NET"


############## start process ########################
table_input = load_workbook(xls_input)
load_value(table_input)
update_table()
table_input.save(xls_output)

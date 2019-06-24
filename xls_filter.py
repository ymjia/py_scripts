# -*- coding: utf-8 -*-
## @file xls_filter.py
## @brief filter error data from xls files
## @author jiayanming

import os.path
import math
import datetime
from openpyxl import load_workbook

class item_m:
    def __init__(self):
        print("item")

    c_id = 0#发票号
    amount = 0 #金额
    b_name = "None" #供应商
    row = -1 #对应行号

def devide_cid(cid):
    if len(cid) == 8:
        return cid

# read valid item from file, record invalid rows
def read_item(file, res_list, invalid_row):
    wb = load_workbook(file)
    ws = wb.active
    for cell in ws.iter_rows(min_row=1, max_col=6, values_only=True):
        print(cell)
    

    
dir_input = "d:/data/xls_f/"
res_list=[1, 2]
invalid_row = [2, 3]
read_item(os.path.join(dir_input, "manual.xlsx"), res_list, invalid_row)
# wb_m = load_workbook(os.path.join(dir_input, "manual.xlsx"))
# wb_d = load_workbook(os.path.join(dir_input, "database.xlsx"))
# ws_m = wb_m.active
# for row in ws_m.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
#     print(row)

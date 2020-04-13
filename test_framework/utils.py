# -*- coding: utf-8 -*-
## @file utils.py
## @brief basic run modules for test framework
## @author jiayanming

import os.path
import sys
import subprocess
import psutil
import time
import threading
import socket
import glob
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis
if sys.platform == "win32":
    import winreg as wr


def get_py_in_reg():
    exe_py = ""
    for reg in [wr.HKEY_CURRENT_USER, wr.HKEY_LOCAL_MACHINE]:
        try:
            reg_table = wr.ConnectRegistry(None, reg)
            for ver in ["3.6", "3.7"]:
                rp = r"Software\Python\PythonCore\3.6\InstallPath".format(ver)
                key = wr.OpenKey(reg_table, rp, 0, wr.KEY_READ)
                try:
                    exe_py, _ = wr.QueryValueEx(key, "ExecutablePath")
                    if exe_py is not None and os.path.exists(exe_py):
                        wr.CloseKey(key)
                        return exe_py
                finally:
                    wr.CloseKey(key)
        except WindowsError:
            pass

def set_reg_item(type_str, val):
    reg = wr.ConnectRegistry(None, wr.HKEY_CURRENT_USER)
    key = None
    try:
        key = wr.CreateKey(reg, r"Software\tf\sts")
    except WindowsError:
        print("cannot find keys in register")
        return
    if key is None:
        return
    wr.SetValueEx(key, type_str, 0, wr.REG_SZ, val)
    wr.CloseKey(key)

def get_reg_item(type_str):
    reg = wr.ConnectRegistry(None, wr.HKEY_CURRENT_USER)
    key = None
    try:
        key = wr.CreateKey(reg, r"Software\tf\sts")
    except WindowsError:
        print("cannot find keys in register")
        return
    if key is None:
        return
    ret = "0"
    try:
        ret = wr.QueryValueEx(key, type_str)[0]
    except WindowsError:
        wr.SetValueEx(key, type_str, 0, wr.REG_SZ, "0")
    wr.CloseKey(key)
    return ret

def get_py_interpretor():
    exe_py = ""
    # findout python interpreter
    if sys.platform == "win32":
        exe_py = get_py_in_reg()
    else:
        exe_rt = os.__file__.split("lib")[0]
        exe_py = os.path.join(exe_rt, "bin", "python3")
    return exe_py


def parse_time(time_str):
    t = time_str.split(" ")
    return t[0], t[1]


def get_latest_file(folder, ext):
    list_f = glob.glob('{}/*.{}'.format(folder, ext))
    if len(list_f) < 1:
        return ""
    return max(list_f, key=os.path.getctime)


def get_sys_table(dir_o, case, ver):
    sys = {}
    dir_log = os.path.join(dir_o, case, ver, "logs")
    file_sys = get_latest_file(dir_log, "sts")
    if not os.path.exists(file_sys):
        print("No log file in {}".format(dir_log))
        return None
    with open(file_sys) as f:
        content = f.readlines()
    str_list = [l.strip() for l in content if len(l) > 4]
    for line in str_list:
        name, v = line.split(" ", 1)
        sys[name] = v
    return sys


# timming table
# read timming info from output/case/version/
def get_time_table(dir_o, case, ver):
    times = {}
    file_time = os.path.join(dir_o, case, ver, "timmings.txt")
    if not os.path.exists(file_time):
        print("{} does not exist".format(file_time))
        return None
    with open(file_time) as f:
        content = f.readlines()
    str_list = [l.strip() for l in content if len(l) > 4]
    for line in str_list:
        name, t = parse_time(line)
        t_flt = 0
        try:
            t_flt = float(t)
        except ValueError:
            continue
        if name in times:
            times[name] += t_flt
        else:
            times[name] = t_flt
    return times


# generate xlsx table file
def get_compare_table(dir_out, l_case, l_ver, ws, table_func):
    #ws = wb.active
    # | alg | case     |    case2   |
    # |     | v1  | v2 |  v1 |  v2  |
    # |alg1 | 0.1 | 0.2|  0.2| 0.3  |
    case_num = len(l_case)
    ver_num = len(l_ver)
    # table title
    title_line = ["Case"]
    ver_line = ["Version"]
    for case in l_case:
        title_line.append(case)
        for i in range(0, ver_num-1):
            title_line.append("")
        for vi in l_ver:
            ver_line.append(vi)
    ws.append(title_line)
    ws.append(ver_line)
    for i in range(0, case_num):
        ws.merge_cells(start_row=1, end_row=1,
                       start_column=i * ver_num + 2,
                       end_column=(i+1) * ver_num + 1)
    # table data
    d_case = {}
    d_ver = {}
    for i in range(0, len(l_case)):
        d_case[l_case[i]] = i
    for i in range(0, len(l_ver)):
        d_ver[l_ver[i]] = i
    # initial quadric item list
    d_alg = {}
    idx_alg = 0
    time_quad = []
    for case in l_case:
        for ver in l_ver:
            #t_alg = get_time_table(dir_out, case, ver)
            t_alg = table_func(dir_out, case, ver)
            if t_alg is None:
                continue
            for alg, t in t_alg.items():
                if alg not in d_alg:
                    d_alg[alg] = idx_alg
                    idx_alg += 1
                time_quad.append([case, ver, alg, t])

    # get cell position for given case,ver,alg
    def get_pos(case, ver, alg):
        if case not in d_case or ver not in d_ver or alg not in d_alg:
            return -1, -1
        row = d_alg[alg] + 3
        cid = d_case[case]
        vid = d_ver[ver]
        col = ver_num * cid + vid + 2
        return row, col
    # fill time data
    for quad in time_quad:
        r, c = get_pos(quad[0], quad[1], quad[2])
        ws.cell(row=r, column=c).value = quad[3]
    # alg name column
    for item in d_alg.items():
        ws.cell(row=item[1] + 3, column=1).value = item[0]
    return 0


support_ext = [".asc", ".rge", ".obj", ".stl", ".ply", ".srge", ".bin"]

def get_sub_dir(folder):
    res = []
    if not os.path.exists(folder):
        return res
    for name in os.listdir(folder):
        if os.path.isdir(os.path.join(folder, name)):
            res.append(name)
    return res


def get_stem_list(folder):
    res = []
    if not os.path.exists(folder):
        return res
    for name in os.listdir(folder):
        if os.path.isdir(os.path.join(folder, name)):
            continue
        stem, ext = os.path.splitext(name)
        if not any(ext in e for e in support_ext):
            continue
        res.append(stem)
    return res

def get_file_list(folder):
    res = []
    if not os.path.exists(folder):
        return res
    for name in os.listdir(folder):
        if os.path.isdir(os.path.join(folder, name)):
            continue
        stem, ext = os.path.splitext(name)
        if not any(ext in e for e in support_ext):
            continue
        res.append(os.path.join(folder, name))
    return res

# def get_file_list(folder):
#     stem_list = get_stem_list(folder)
#     res = []
#     for stem in stem_list:
#         res.append(os.path.join(folder, stem))
#     return res


def get_selected_item(qlv):
    res = []
    sl = qlv.selectedIndexes()
    if len(sl) < 1:
        return res
    for s in sl:
        res.push(s.data())


def get_sys_info():
    res = []
    res.append("CPU_freq {}MHz\n".format(psutil.cpu_freq().max))
    res.append("CPU_core {}\n".format(psutil.cpu_count(logical=False)))
    res.append("CPU_thread {}\n".format(psutil.cpu_count()))
    res.append("MEM_total {0:.2f}MB\n".format(psutil.virtual_memory().total / 1e6))
    res.append("MEM_virtual {0:.2f}MB\n".format(psutil.swap_memory().total / 1e6))
    net_info = psutil.net_if_addrs()
    net_id = 0
    for net in net_info:
        cur_net = net_info[net]
        for item in cur_net:
            if item.family != socket.AF_INET:
                continue
            res.append("PC_ip_{} {}\n".format(net_id, item.address))
            net_id += 1
    res.append("USER {}\n".format(os.getlogin()))
    res.append("Platform {}\n".format(sys.platform))
    return res


def background_monitor(pm):
    while pm.poll():
        time.sleep(.3)


class ProcessMonitor:
    def __init__(self, command, f_log):
        self.command = command
        self.execution_state = False
        self._fLog = f_log
        self.p = None
        self._cpuSample = []
        self._memSample = []

    def execute(self):
        self.max_vmem = 0
        self.max_pmem = 0
        self.t1 = None
        self.t0 = time.time()
        if len(self.command) < 1:
            return
        dir_exe = os.path.dirname(self.command[0])
        try:
            self.p = psutil.Popen(
                self.command, shell=False, cwd=dir_exe,
                stdout=subprocess.PIPE, stderr=subprocess.STDOUT)
        except PermissionError:
            print("Running Error in: {}".format(" ".join(self.command)))
            return False
        self.execution_state = True
        return True

    def monite_execution(self):
        t = threading.Thread(target=background_monitor, args=(self,), daemon=True)
        t.start()

    def poll(self):
        if not self.check_execution_state():
            return False
        self.t1 = time.time()
        try:
            pp = psutil.Process(self.p.pid)
            #time.sleep(0.1)
            # obtain a list of the subprocess and all its descendants
            descendants = list(pp.children(recursive=True))
            descendants = descendants + [pp]

            rss_memory = 0
            vms_memory = 0

            #calculate and sum up the memory of the subprocess and all its descendants 
            for descendant in descendants:
                try:
                    mem_info = descendant.memory_info()
                    rss_memory += mem_info[0]
                    vms_memory += mem_info[1]
                except psutil.NoSuchProcess:
                    #sometimes a subprocess descendant will have terminated between the tim
                    # we obtain a list of descendants, and the time we actually poll this
                    # descendant's memory usage.
                    pass
            self.max_vmem = max(self.max_vmem, vms_memory)
            self.max_pmem = max(self.max_pmem, rss_memory)
            self._memSample.append(rss_memory / 1e6)
            self._cpuSample.append(pp.cpu_percent(interval=1) / psutil.cpu_count())
        except psutil.NoSuchProcess:
            return self.check_execution_state()
        return self.check_execution_state()

    def is_running(self):
        return psutil.pid_exists(self.p.pid) and self.p.poll() is None

    def check_execution_state(self):
        if not self.execution_state:
            return False
        if self.is_running():
            return True
        self.executation_state = False
        self.t1 = time.time()
        return False

    def close(self, kill=False):
        if self.p is None:
            return
        try:
            pp = psutil.Process(self.p.pid)
            if kill:
                pp.kill()
            else:
                pp.terminate()
        except psutil.NoSuchProcess:
            pass


# class for global config
## read config file, generate information for test framework
## @brief configuration for current session
## @note for information transfer from ui to pvpython.exe
class SessionConfig:
    def __init__(self):
        self.list_case = []
        self.list_ver = []
        self.list_alg = []
        self.config_map = {}
        # parameter for hausdorff
        self.config_map["hd_critical_dist"] = "0.05"
        self.config_map["hd_nominal_dist"] = "0.03"
        self.config_map["hd_max_dist"] = "0.3"
        self.config_map["hd_single_color"] = "True"
        # parameter for screenshot
        self.config_map["ss_force_update"] = "False"
        self.config_map["rep_specular"] = "True"
        self.config_map["view_width"] = "1024"
        self.config_map["view_height"] = "768"
        self.config_map["transparent_background"] = "False"

    def config_val(self, key_str, default_val):
        if key_str in self.config_map:
            return self.config_map[key_str]
        return default_val
        
    def read_config(self, filename):
        if not os.path.exists(filename):
            print("Warning! config file {} does not exists".format(filename))
            return False
        content = None
        with open(filename, encoding='utf-8') as f:
            content = f.readlines()
        lines = [l.strip() for l in content]
        if len(lines) < 3:
            print("Warning! Invalid config file {}".format(filename))
            return False
        # fix part
        if lines[0][0:3] != "cas" or lines[1][0:3] != "ver" or lines[2][0:3] != "alg":
            print("Warning! Invalid config file {}".format(filename))
            return False
        self.list_case = lines[0].split(" ")[1:].copy()
        self.list_ver = lines[1].split(" ")[1:].copy()
        self.list_alg = lines[2].split(" ")[1:].copy()
        # optional part
        for i in range(4, len(lines)):
            l_couple = lines[i].split(" ")
            if len(l_couple) != 2:
                continue
            self.config_map[l_couple[0]] = l_couple[1]
        return True

    def write_config(self, filename):
        f_config = open(filename, "w", encoding='utf-8')
        f_config.writelines("cas {}\n".format(" ".join(map(str, self.list_case))))
        f_config.writelines("ver {}\n".format(" ".join(map(str, self.list_ver))))
        f_config.writelines("alg {}\n".format(" ".join(map(str, self.list_alg))))
        for key, val in self.config_map.items():
            f_config.writelines("{} {}\n".format(key, val))
        f_config.close()


# indent xml ElementTree.root
def indent_xml(elem, level=0):
    i = "\n" + level*"  "
    if len(elem):
        if not elem.text or not elem.text.strip():
            elem.text = i + "  "
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
        for elem in elem:
            indent_xml(elem, level+1)
        if not elem.tail or not elem.tail.strip():
            elem.tail = i
    else:
        if level and (not elem.tail or not elem.tail.strip()):
            elem.tail = i


def create_f_lists(dir_o, l_case, l_ver, f_list, sp_list):
    for case in l_case:
        for ver in l_ver:
            dir_log = os.path.join(dir_o, case, ver, "logs")
            file_sys = get_latest_file(dir_log, "smp")
            if not os.path.exists(file_sys):
                continue
            f_list.append(file_sys)
            sp_list.append("{}_{}".format(case, ver))


# read in list files, create chart on f_out
# sp_list: column name
def create_chart(in_list, sp_list, ws):
    f_num = len(in_list)
    if f_num < 1:
        return 0
    # prepare data
    mem_smp = []
    cpu_smp = []
    for f_in in in_list:
        with open(f_in) as f:
            content = f.readlines()
        cur_mem = []
        cur_cpu = []
        str_list = [l.strip() for l in content if len(l) > 4]
        for line in str_list:
            cpu, mem = line.split(" ", 1)
            cur_mem.append(mem)
            cur_cpu.append(cpu)
        mem_smp.append(cur_mem)
        cpu_smp.append(cur_cpu)
    # create xlsx data table
    max_row = 0
    for f in range(0, f_num):
        cur_mem = mem_smp[f]
        cur_cpu = cpu_smp[f]
        c = f + 1
        r_num = len(cur_cpu)
        if r_num > max_row:  # assersion: len(cpu) equal to len(mem)
            max_row = r_num
        title_cpu = "cpu"
        title_mem = "mem"
        if len(sp_list) > f:
            title_cpu = "{}_cpu".format(sp_list[f])
            title_mem = "{}_mem".format(sp_list[f])
        ws.cell(row=1, column=f + 1).value = title_cpu
        ws.cell(row=1, column=f + f_num + 1).value = title_mem
        for r in range(0, r_num):
            ws.cell(row=r+2, column=c).value = float(cur_cpu[r])
        c = f + f_num + 1
        for r in range(0, len(cur_mem)):
            ws.cell(row=r+2, column=c).value = float(cur_mem[r])
    # create chart
    lc_cpu = LineChart()
    data_cpu = Reference(ws, min_col=1, min_row=1,
                         max_col=f_num, max_row=max_row + 1)
    lc_cpu.add_data(data_cpu, titles_from_data=True)
    lc_cpu.title = "CPU/Mem Usage Line"
    lc_cpu.style = 12
    lc_cpu.y_axis.title = "CPU"
    lc_cpu.x_axis.title = "Time"
    lc_cpu.y_axis.crosses = "max"  # for merge chart
    # Create a second chart
    lc_mem = LineChart()
    data_mem = Reference(ws, min_col=f_num + 1, min_row=1,
                         max_col=f_num * 2, max_row=max_row + 1)
    lc_mem.add_data(data_mem, titles_from_data=True)
    lc_mem.y_axis.axId = 200
    lc_mem.y_axis.title = "MEM"
    lc_mem.y_axis.majorGridlines = None
    for se in lc_mem.series:
        se.graphicalProperties.line.dashStyle = "sysDot"
    lc_cpu += lc_mem
    ws.add_chart(lc_cpu, "{}1".format(chr(ord('A') + f_num * 2 + 2)))
    return 0


if __name__ == "__main__":
    # in_list = ["c:/data/test_framework/management/project1/output/case1/test/logs/tfl_20190820_095928.smp", "c:/data/test_framework/management/project1/output/case1/test/logs/tfl_20190820_095204.smp"]
    # create_chart(in_list, "c:/tmp/res.xlsx", ["case1", "case2"])
    # os.startfile("c:/tmp/res.xlsx")
    set_reg_item("exe", "5")
    print(get_reg_item("ss"))
 

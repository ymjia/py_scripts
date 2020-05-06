# -*- coding: utf-8 -*-
## @file ui_logic.py
## @brief ui signal/slot logitcs
## @author jiayanming

import os.path
import sys
import datetime
import subprocess
import xml.etree.ElementTree as ET
from openpyxl import Workbook

from PyQt5.QtWidgets import (QApplication, QListView, QFileDialog, QMessageBox,
                             QInputDialog, QLineEdit, QPushButton, QProgressDialog,
                             QAbstractItemView)
from PyQt5.QtCore import QItemSelectionModel, Qt
from PyQt5.QtGui import QStandardItemModel, QStandardItem

from test_framework import project_io
from test_framework.project_io import get_checked_items
from test_framework import generate_docx
from test_framework import utils
from test_framework import thread_module
from test_framework import create_screenshots
from paraview.simple import Interact, CreateView, Render, Delete

FILEBROWSER = "explorer"
if sys.platform != "win32":
    FILEBROWSER = "open" if sys.platform == "darwin" else "xdg-open"


# open file for mac
def open_file(filename):
    if not os.path.exists(filename):
        print("Warning! {} not exists!".format(filename))
    if sys.platform == "win32":
        os.startfile(filename)
    else:
        subprocess.call([FILEBROWSER, filename])


def explore(path):
    path = os.path.normpath(path)
    if os.path.isdir(path):
        open_file(path)
    elif os.path.isfile(path):
        if sys.platform == "win32":
            subprocess.call([FILEBROWSER, '/select,', path])
        else:
            subprocess.call([FILEBROWSER, os.path.dirname(path)])


def ptree_add_item(pt, path):
    stem = os.path.splitext(os.path.basename(path))[0]
    root = pt.getroot()
    for item in root:
        if item.attrib["name"] == stem:
            return item
    new_item = ET.Element("item", {"name": stem, "path": path})
    root.append(new_item)
    return new_item


def get_qlist_idx(qlv, name):
    model = qlv.model()
    for index in range(model.rowCount()):
        if model.item(index).text() == name:
            return index
    return -1


def set_project_selected(qlv, name):
    row = get_qlist_idx(qlv, name)
    if row < 0:
        return
    q_idx = qlv.model().index(row, 0)
    qlv.selectionModel().select(q_idx, QItemSelectionModel.Select)


def get_selected_project_item(ui):
    sl = ui._qlv_all_proj.selectedIndexes()
    if len(sl) < 1:
        return None
    return find_ptree_item(ui._pTree, sl[0].data())


def find_ptree_item(pt, name):
    root = pt.getroot()
    for item in root:
        if item.attrib["name"] == name:
            return item
    return None

def slot_generate_docx(ui):
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    doc_type = p_obj._curDocType
    if doc_type == "Screenshots":
        return generate_ss_docx(ui)
    elif doc_type == "Time_statistics":
        return generate_time_docx(ui)
    elif doc_type == "CPU_MEM_statistics":
        return generate_proc_docx(ui)
    else:
        return generate_hausdorf_docx(ui)


def generate_ss_docx(ui):
    p_obj = ui._p
    dir_i = p_obj._dirInput
    dir_o = p_obj._dirOutput
    dir_doc = os.path.join(dir_o, "doc")
    doc_name = p_obj._docName
    str_time = str(datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
    doc_name_final = "{}_{}.docx".format(doc_name, str_time)
    p_obj._curDocName = doc_name_final
    if not os.path.exists(dir_doc):
        os.makedirs(dir_doc)
    file_save = os.path.join(dir_doc, doc_name_final)
    l_case = get_checked_items(p_obj._case, p_obj._dCaseCheck)
    l_ver = get_checked_items(p_obj._ver, p_obj._dVerCheck)
    l_alg = get_checked_items(p_obj._alg, p_obj._dAlgCheck)
    gd = generate_docx.DocxGenerator(dir_i, dir_o, l_case, l_ver, l_alg)
    gd.generate_docx(file_save, p_obj._configFile)
    ui.add_hist_item("doc", 1)
    QMessageBox.about(ui, "Message", "Docx wrote to {}!".format(file_save))


def generate_time_docx(ui):
    p_obj = ui._p
    dir_o = p_obj._dirOutput
    dir_doc = os.path.join(dir_o, "doc")
    tdoc_name = p_obj._docName + "_time"
    str_time = str(datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
    tdoc_name_final = "{}_{}.xlsx".format(tdoc_name, str_time)
    p_obj._curDocName = tdoc_name_final
    if not os.path.exists(dir_doc):
        os.makedirs(dir_doc)
    file_save = os.path.join(dir_doc, tdoc_name_final)
    l_case = get_checked_items(p_obj._case, p_obj._dCaseCheck)
    l_ver = get_checked_items(p_obj._ver, p_obj._dVerCheck)
    wb = Workbook()
    ws = wb.active
    res = utils.get_compare_table(dir_o, l_case, l_ver, ws, utils.get_time_table)
    try:
        wb.save(file_save)
    except PermissionError:
        res = 1
    if res == 0:
        ui.add_hist_item("doc", 1)
        QMessageBox.about(ui, "Message", "Docx wrote to {}!".format(file_save))
    else:
        QMessageBox.about(ui, "Error", "Cannot wrote to {}!".format(file_save))


def generate_proc_docx(ui):
    p_obj = ui._p
    dir_o = p_obj._dirOutput
    dir_doc = os.path.join(dir_o, "doc")
    pdoc_name = p_obj._docName + "_proc"
    str_time = str(datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
    pdoc_name_final = "{}_{}.xlsx".format(pdoc_name, str_time)
    p_obj._curDocName = pdoc_name_final
    if not os.path.exists(dir_doc):
        os.makedirs(dir_doc)
    file_save = os.path.join(dir_doc, pdoc_name_final)
    l_case = get_checked_items(p_obj._case, p_obj._dCaseCheck)
    l_ver = get_checked_items(p_obj._ver, p_obj._dVerCheck)
    wb = Workbook()
    ws = wb.active
    res = utils.get_compare_table(dir_o, l_case, l_ver, ws, utils.get_sys_table)
    ws_chart = wb.create_sheet("chart", 0)
    f_list = []
    sp_list = []
    utils.create_f_lists(dir_o, l_case, l_ver, f_list, sp_list)
    utils.create_chart(f_list, sp_list, ws_chart)
    try:
        wb.save(file_save)
    except PermissionError:
        res = 1
    if res == 0:
        ui.add_hist_item("doc", 1)
        QMessageBox.about(ui, "Message", "Docx wrote to {}!".format(file_save))
    else:
        QMessageBox.about(ui, "Error", "Cannot wrote to {}!".format(file_save))


def generate_hausdorf_docx(ui):
    # create hausdorff shot
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    sc = utils.SessionConfig()
    sc.dir_i = p_obj._dirInput
    sc.dir_o = p_obj._dirOutput
    sc.list_case = get_checked_items(p_obj._case, p_obj._dCaseCheck)
    sc.list_ver = ['__hausdorff']
    sc.list_alg.clear()
    total_num = create_screenshots.create_hausdorff_shot(sc)
    if total_num > 0:
        ui.add_hist_item("ss", total_num)
    print(total_num)
    # generate doc
    doc_name = p_obj._docName
    str_time = str(datetime.datetime.now().strftime("%Y%m%d_%H%M%S"))
    doc_name_final = "{}_{}_hd.docx".format(doc_name, str_time)
    p_obj._curDocName = doc_name_final
    dir_doc = os.path.join(p_obj._dirOutput, "doc")
    if not os.path.exists(dir_doc):
        os.makedirs(dir_doc)
    file_save = os.path.join(dir_doc, doc_name_final)
    l_ver = ["hausdorff_A2B", "hausdorff_B2A"]
    l_alg = ["__hd"] # reserved alg_name for hausdorff dist
    gd = generate_docx.DocxGenerator(p_obj._dirInput, p_obj._dirOutput, sc.list_case,
                                     l_ver, l_alg)
    gd.generate_docx(file_save, p_obj._configFile)
    ui.add_hist_item("doc", 1)
    QMessageBox.about(ui, "Message", "Hausdorff Docx wrote to {}!".format(file_save))
    

def slot_open_docx(ui):
    p_obj = ui._p
    dir_doc = os.path.join(p_obj._dirOutput, "doc")
    file_doc = os.path.join(dir_doc, p_obj._curDocName)
    if os.path.exists(file_doc):
        open_file(file_doc)
    else:
        QMessageBox.about(ui, "Error", "Document doesnot Exist! Try Generate Docx First.")


def slot_open_docx_path(ui):
    p_obj = ui._p
    dir_doc = os.path.join(p_obj._dirOutput, "doc")
    if not os.path.exists(dir_doc):
        os.makedirs(dir_doc)
    open_file(dir_doc)


def slot_create_screenshots(ui):
    qm = QMessageBox
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    # create session config object
    sc = utils.SessionConfig()
    sc.list_case = get_checked_items(p_obj._case, p_obj._sCaseCheck)
    sc.list_alg = get_checked_items(p_obj._alg, p_obj._sAlgCheck)
    sc.list_ver = get_checked_items(p_obj._ver, p_obj._sVerCheck)
    sc.dir_i = p_obj._dirInput
    sc.dir_o = p_obj._dirOutput

    if len(sc.list_case) < 1:
        QMessageBox.about(ui, "Error", "No Case checked")
        return
    if len(sc.list_ver) < 1:
        QMessageBox.about(ui, "Error", "No Version checked")
        return
    if len(sc.list_alg) < 1:
        ret = qm.question(ui, "", "No FileNames checked, Continue?", qm.Yes | qm.Cancel)
        if ret == qm.Cancel:
            return
    pg = QProgressDialog("Creating ScreenShots...", "Cancel", 0, 100, ui)
    pg.setWindowModality(Qt.WindowModal)
    pg.setCancelButton(None)
    pg.setValue(1)
    pg.show()
    total_num = create_screenshots.create_screenshots(sc, pg.setValue)
    if total_num > 0:
        ui.add_hist_item("ss", total_num)
    QMessageBox.about(ui, "Message", "Create Screenshots Completed! {} file generated".format(total_num))


def get_default_path(in_path):
    if in_path is None:
        return "c:/"
    d = in_path
    while not os.path.exists(d):
        pd = os.path.dirname(d)
        if pd == d or len(pd) < 1:
            break
        d = pd
    if not os.path.exists(d):
        d = "c:/"
    return d


def get_default_proj_path(ui):
    p_obj = ui._p
    if p_obj is not None:
        config_file = p_obj._configFile
        if os.path.exists(config_file):
            return get_default_path(config_file)
    # get default from p_list
    if ui._pTree is not None:
        for item in ui._pTree.getroot():
            p_path = item.attrib["path"]
            if os.path.exists(p_path):
                return get_default_path(p_path)
    # return project path
    return get_default_path(ui._ptName)


# get file and set qle.text
def slot_get_path(qle):
    d = get_default_path(qle.text())
    path = QFileDialog.getExistingDirectory(None, 'Select a folder:', d, QFileDialog.ShowDirsOnly)
    if path is not None and path != "":
        qle.setText(path)


# open path in qle.text
def slot_open_path(qle):
    p = qle.text()
    if not os.path.exists(p):
        QMessageBox.about(None, "Error", "{} does not exist!".format(p))
        return
    explore(p)


def slot_get_file(qle, f_type):
    path = ""
    d = get_default_path(qle.text())
    if f_type == "xml":
        path, _filter = QFileDialog.getOpenFileName(None, 'Open File', d, 'XML (*.xml)')
    else:
        path, _filter = QFileDialog.getOpenFileName(None, 'Open File', d, 'EXE (*.exe);;PY (*.py);;All Files (*)')
    if path is not None and path != "":
        qle.setText(path)


def slot_open_input_path(ui):
    in_depth = int(utils.g_config.config_val("exe_input_depth", "1"))
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    d = get_default_path(p_obj._dirInput)
    dir_in = QFileDialog.getExistingDirectory(None, 'Select a folder:', d, QFileDialog.ShowDirsOnly)
    if dir_in is None or dir_in == "":
        return
    p_obj._dirInput = dir_in
    # p_obj._case = [f for f in os.listdir(dir_in)
    #                if os.path.isdir(os.path.join(dir_in, f))]
    p_obj._case = utils.list_subdir_with_depth(dir_in, in_depth)
    p_obj._eCaseCheck.clear()
    p_obj._sCaseCheck.clear()
    p_obj._dCaseCheck.clear()
    ui.fill_ui_info(p_obj)


def replace_sep(in_str):
    if sys.platform == "win32":
        return in_str.replace("/", "\\").replace("\\\\", "\\")
    else:
        return in_str.replace("\\", "/").replace("//", "/")


def generate_exe_param(dir_i, dir_o, case, exe, cmd, ver):
    org_param = cmd
    p_i = ""
    # find input place holder
    ph_i = "{i}" # place holder for input
    auto_input = utils.g_config.config_val("exe_auto_input", True)
    i_pos = int(org_param.find(ph_i))
    p_len = len(org_param)
    if i_pos != -1:
        # find string follows {i} until meet 'space'
        i_end = i_pos + 3
        while i_end < p_len and org_param[i_end] != ' ':
            i_end += 1
        i_suffix = org_param[i_pos + 3: i_end]
        ph_i = ph_i + i_suffix
        p_i = os.path.join(dir_i, case) + i_suffix
        if not os.path.exists(p_i):
            print("Error! input {} does not Exist!".format(p_i))
            return ""
        # use file name if only one file in case dir
        if auto_input and not os.path.isfile(p_i):
            i_list = utils.get_file_list(p_i)
            if len(i_list) == 1:
                p_i = i_list[0]
        p_i = replace_sep(p_i)
    p_o = ""
    if "{o}" in org_param:
        # get output param
        p_o = os.path.join(dir_o, case, ver) + "/"
        p_o = replace_sep(p_o)
    # replace
    param = org_param.replace(ph_i, p_i).replace("{o}", p_o).replace("{c}", case).replace("{v}", ver)
    return param


def slot_exe_run(ui):
    ui._p = ui.collect_ui_info()
    ui._p.save_xml(ui._p._configFile)
    p_obj = ui._p

    exe = p_obj._exeDemo
    param_text = p_obj._exeParam
    if not os.path.exists(exe):
        QMessageBox.about(ui, "Error", "Demo {} does not exist!".format(exe))
        return
    list_case = get_checked_items(p_obj._case, p_obj._eCaseCheck)
    if len(list_case) < 1 and "{i}" in param_text:
        QMessageBox.about(ui, "Error", "No Input Case Checked!!")
        return

    ss_exe = thread_module.ExeSession(ui._p, False)
    
    ui._cmdDialog.add_cmd([[exe, param_text]])
    ui._threadExe = thread_module.ExeRunThread(ss_exe)
    ui._threadExe.setTerminationEnabled()
    ui._threadExe._sigProgress.connect(ui.exe_progress)
    ui._threadExe.finished.connect(lambda: ui.exe_finish(False))
    ui.new_stop_button()
    ui._qlv_all_proj.setDisabled(True)
    ui._threadExe.start()


def slot_exe_stop(ui):
    if ui._threadExe is not None:
        if ui._threadExe._demoProc is not None:
            ui._threadExe._demoProc.close(kill=True)
        ui._threadExe.release_files()
        ui._threadExe.terminate()
        ui._threadExe.wait()
        QMessageBox.about(ui, "Message", "Demo Terminated!")


def slot_exe_param(ui):
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    dir_i = p_obj._dirInput
    dir_o = p_obj._dirOutput
    exe = ui._p._exeDemo
    cmd = ui._p._exeParam
    ver = ui._p._eVer
    list_case = get_checked_items(p_obj._case, p_obj._eCaseCheck)
    if len(list_case) < 1:
        QMessageBox.about(ui, "Error", "No CheckItem in Input Case!")
        return
    param_list = "ParamLine Preview:"
    for case in list_case:
        param_list += "\n\n"
        param_list += generate_exe_param(dir_i, dir_o, case, exe, cmd, ver)
    QMessageBox.about(ui, "Message", param_list)


def slot_new_project(ui):
    dp = get_default_proj_path(ui)
    path = ""
    while True:
        path, _ = QFileDialog.getSaveFileName(None, "Save New Project", dp, "XML(*.xml)")
        if path is None or path == "":
            return
        base_filename = os.path.basename(path)
        stem, ext = os.path.splitext(base_filename)
        if stem == "tf_proj" or stem == "cmd_history":
            QMessageBox.about(ui, "Error", "'tf_proj' and 'cmd_history' is reserved filename, please change project name!")
            continue
        if ext != ".xml":
            path += ".xml"
        break
    if path == "":
        return
    p = project_io.Project()
    p._configFile = path
    ui._p = p
    ui._p.save_xml(path)
    ui.fill_ui_info(ui._p)
    new_item = ptree_add_item(ui._pTree, path)
    ui.fill_proj_list()
    set_project_selected(ui._qlv_all_proj, new_item.attrib["name"])
    utils.indent_xml(ui._pTree.getroot())
    ui._pTree.write(ui._ptName)


def slot_delete_project(ui):
    sl = ui._qlv_all_proj.selectedIndexes()
    if len(sl) < 1:
        QMessageBox.about(ui, "Error", "No Selection to Delete!")
        return
    # save old
    del_item = find_ptree_item(ui._pTree, sl[0].data())
    ui._pTree.getroot().remove(del_item)
    utils.indent_xml(ui._pTree.getroot())
    ui._pTree.write(ui._ptName)
    ui.fill_proj_list()
    # get new
    new_p = project_io.Project()
    ui._p = new_p
    ui.fill_ui_info(ui._p)
    return


def slot_save_project(ui):
    # update project object
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    # update xml object
    p_obj._tree = p_obj.collect_xml()
    # save xml
    p_obj.save_xml()


def slot_load_project(ui):
    dp = get_default_proj_path(ui)
    path, _filter = QFileDialog.getOpenFileName(None, 'Open File', dp, 'XML (*.xml)')
    if path is None or path == "":
        return
    if not os.path.exists(path):
        QMessageBox.about(None, "Error", "Wrong file: {}!".format(path))
        return
    p = project_io.Project()
    p.load_xml(path)
    ui.fill_ui_info(p)
    new_item = ptree_add_item(ui._pTree, path)
    ui.fill_proj_list()
    set_project_selected(ui._qlv_all_proj, new_item.attrib["name"])
    utils.indent_xml(ui._pTree.getroot())
    ui._pTree.write(ui._ptName)
    return


def slot_scan_input(ui):
    in_depth = int(utils.g_config.config_val("exe_input_depth", "1"))
    # update project_object
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    dir_in = p_obj._dirInput
    p_obj._case = utils.list_subdir_with_depth(dir_in, in_depth)
    # p_obj._case = [f for f in os.listdir(dir_in)
    #                if os.path.isdir(os.path.join(dir_in, f))]
    ui.fill_ui_info(p_obj)
    return


# build dir for all case/version
def slot_build_output(ui):
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    dir_out = p_obj._dirOutput
    l_case = get_checked_items(p_obj._case, p_obj._sCaseCheck)
    l_ver = get_checked_items(p_obj._ver, p_obj._sVerCheck)
    for c in l_case:
        for v in l_ver:
            n_dir = os.path.join(dir_out, c, v)
            if os.path.exists(n_dir):
                continue
            os.makedirs(n_dir)


def append_input_to_list(l, label):
    text, ok = QInputDialog.getText(None, label, "Item list, seperate with ',':", QLineEdit.Normal, "")
    if ok and text != '':
        item_list = text.replace(" ", "").split(",")
        for i in item_list:
            if i in l:
                continue
            l.append(i)
        return 0
    return 1


def slot_add_list(ui, obj_list, type_str):
    ui._p = ui.collect_ui_info()
    if append_input_to_list(obj_list, type_str) == 0:
        ui.fill_ui_info(ui._p)
    return


def del_list_item(ui, qlv, item_list):
    sl = qlv.selectedIndexes()
    if len(sl) < 1:
        QMessageBox.about(None, "Error", "No Selection to Delete!")
        return
    for s in sl:
        item_name = s.data()
        for i in range(0, len(item_list)):
            if item_list[i] == item_name:
                item_list.pop(i)
                break


def slot_del_list(ui, qlv, obj_list):
    ui._p = ui.collect_ui_info()
    del_list_item(ui, qlv, obj_list)
    ui.fill_ui_info(ui._p)
    return


def slot_qlv_double_click(ui, qlv, qle):
    sl = qlv.selectedIndexes()
    if len(sl) < 1:
        return
    click_path = os.path.join(qle.text(), sl[0].data())
    if os.path.exists(click_path):
        explore(click_path)
    return


def slot_ss_manage(ui):
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    sl = ui._qlv_ss_case.selectedIndexes()
    if len(sl) < 1:
        QMessageBox.about(None, "Tip:", "No Selected Item in CASE LIST to Manage!")
        return
    case_name = sl[0].data()
    dir_case = os.path.join(p_obj._dirInput, case_name)
    dir_case_out= os.path.join(p_obj._dirOutput, case_name)
    if not os.path.exists(dir_case_out):
        os.makedirs(dir_case_out)
    # default directory show in file dialog
    dir_default = dir_case
    if utils.g_config.config_val("ss_default_reference_directory", "Input") == "Output":
        dir_default = dir_case_out

    file_config = os.path.join(dir_case, "config.txt")
    if not os.path.exists(file_config):
        f= open(file_config, "w+")
        f.close()

    # popup a window to select file,
    # if not exist or not readable, use old version
    # else read file and render window
    type_filter = "Models (*.stl *.ply *.obj *.asc *.rge *.tb)"
    fo = QFileDialog()
    #fo.setOption(QFileDialog.DontUseNativeDialog, False)
    #fo.setHistory([dir_case, dir_case_out])
    fo.setFileMode(QFileDialog.ExistingFiles)
    names, _ = fo.getOpenFileNames(ui, "Select Model", dir_default, type_filter)
    # camera already set
    if len(names) > 0:
        create_screenshots.start_camera_set_session(names, file_config)
        return
    open_file(dir_case)


def slot_ss_preview(ui):
    ui._p = ui.collect_ui_info()
    p_obj = ui._p
    # get dir
    sl = ui._qlv_ss_case.selectedIndexes()
    if len(sl) < 1:
        QMessageBox.about(None, "Tip:", "No Selected Item in CASE LIST to Manage!")
        return
    case_name = sl[0].data()
    sc = utils.SessionConfig()
    dir_i = p_obj._dirInput
    dir_o = p_obj._dirOutput
    sc.dir_i = dir_i
    sc.dir_o = dir_o
    sc.list_case = [case_name]
    sc.list_ver = ["input"]
    sc.list_alg = ["tmp"]
    total_num = create_screenshots.create_screenshots(sc)
    # write to file
    first_pic = os.path.join(dir_o, case_name, "input")
    if os.path.exists(first_pic):
        open_file(first_pic)
    return


def load_ptree_obj(ui):
    dir_lp = os.path.dirname(os.path.realpath(__file__))
    file_lp = os.path.join(dir_lp, "tf_proj.xml")
    if not os.path.exists(file_lp):
        # create new xml
        if not os.path.exists(dir_lp):
            os.makedirs(dir_lp)
        root_new = ET.Element("projects")
        ui._pTree = ET.ElementTree(root_new)
        utils.indent_xml(ui._pTree.getroot())
        ui._pTree.write(file_lp)
    else:
        # has existing, load info
        ui._pTree = ET.parse(file_lp)


def save_ptree_obj(ui):
    dir_lp = os.path.dirname(os.path.realpath(__file__))
    file_lp = os.path.join(dir_lp, "tf_proj.xml")
    utils.indent_xml(ui._pTree.getroot())
    ui._pTree.write(file_lp)
    return


def slot_switch_proj(ui):
    if not ui._qlv_all_proj.isEnabled():
        return
    sl = ui._qlv_all_proj.selectedIndexes()
    if len(sl) < 1:
        return
    # save old
    ui._p = ui.collect_ui_info()
    ui._p.save_xml(ui._p._configFile)
    # get new
    new_item = find_ptree_item(ui._pTree, sl[0].data())
    if new_item is None:
        return
    new_xml = new_item.attrib["path"]
    if not os.path.exists(new_xml):
        QMessageBox.about(None, "Error", "{} doesnot exist!".format(new_xml))
        return
    new_p = project_io.Project()
    new_p.load_xml(new_xml)
    ui._p = new_p
    ui.fill_ui_info(ui._p)
    ui._pTree.write(ui._ptName)


def slot_open_proj_path(ui):
    sl = ui._qlv_all_proj.selectedIndexes()
    if len(sl) < 1:
        return
    click_path = ui._p._configFile
    if not os.path.exists(click_path):
        QMessageBox.about(None, "Error", "{} doesnot exist!".format(click_path))
        return
    explore(click_path)

def get_ver_dir(dir_o, qlv_case, qlv_ver):
    csl = qlv_case.selectedIndexes()
    if len(csl) < 1:
        return "1"
    vsl = qlv_ver.selectedIndexes()
    if len(vsl) < 1:
        return "2"
    v_dir = os.path.join(dir_o, csl[0].data(), vsl[0].data())
    return v_dir


def get_alg_file_name(dir_o, qlv_case, qlv_ver, qlv_alg):
    v_dir = get_ver_dir(dir_o, qlv_case, qlv_ver)
    if len(v_dir) < 3:
        return v_dir
    if not os.path.exists(v_dir):
        return "4"
    asl = qlv_alg.selectedIndexes()
    if len(asl) < 1:
        return "3"
    stem = asl[0].data()
    alg_file = os.path.join(v_dir, stem)
    if os.path.exists(alg_file):
        # alg output is directory
        return alg_file
    # find file with alg stem
    for f in os.listdir(v_dir):
        cur_stem = os.path.splitext(f)[0]
        if cur_stem == stem:
            return os.path.join(v_dir, f)
    return "5"


def print_list_error_message(err_str):
    if err_str == "1":
        QMessageBox.about(None, "Error", "No Case slected in Case List!")
        return
    if err_str == "2":
        QMessageBox.about(None, "Error", "No Version slected in Version List!")
        return
    if err_str == "3":
        QMessageBox.about(None, "Error", "No FileName slected in FileName List!")
        return
    if err_str == "4":
        QMessageBox.about(None, "Error", "No Version folder in Output Dir!")
        return
    if err_str == "5":
        QMessageBox.about(None, "Error", "Cannot Find FileName in Output Dir!")
        return
    QMessageBox.about(None, "Error", "{} does not exists!".format(err_str))


def slot_open_ss_ver(ui):
    res = get_ver_dir(ui._p._dirOutput, ui._qlv_ss_case, ui._qlv_ss_ver)
    if not os.path.exists(res):
        print_list_error_message(res)
        return
    explore(res)


def slot_open_ss_alg(ui):
    res = get_alg_file_name(ui._p._dirOutput, ui._qlv_ss_case, ui._qlv_ss_ver, ui._qlv_ss_alg)
    if not os.path.exists(res):
        print_list_error_message(res)
        return
    explore(res)


def slot_open_doc_ver(ui):
    res = get_ver_dir(ui._p._dirOutput, ui._qlv_doc_case, ui._qlv_doc_ver)
    if not os.path.exists(res):
        print_list_error_message(res)
        return
    explore(res)


def slot_open_doc_alg(ui):
    res = get_alg_file_name(ui._p._dirOutput, ui._qlv_doc_case, ui._qlv_doc_ver, ui._qlv_doc_alg)
    if not os.path.exists(res):
        print_list_error_message(res)
        return
    explore(res)


def create_QListView(ui, qle=None):
    ql = QListView(ui)
    ql.setEditTriggers(QAbstractItemView.NoEditTriggers)
    ql.setSelectionMode(QAbstractItemView.ExtendedSelection)
    ql.setDefaultDropAction(Qt.TargetMoveAction)
    ql.setDragDropMode(QAbstractItemView.InternalMove);
    ql.setMinimumHeight(80)
    ql.clicked.connect(lambda: slot_qlv_check_list(ui, ql))
    if qle is not None:
        ql.doubleClicked.connect(lambda: slot_qlv_double_click(ui, ql, qle))
    return ql

# get listview from project_object
def fill_check_list(lv, item_list, check_dict):
    model = QStandardItemModel()
    #pt_item = QStandardItem()
    flag = Qt.ItemIsSelectable | Qt.ItemIsDragEnabled | Qt.ItemIsEnabled
    for i in item_list:
        item = QStandardItem(i)
        item.setFlags(flag)
        check = Qt.Checked if i in check_dict else Qt.Unchecked
        item.setCheckState(check)
        item.setCheckable(True)
        model.appendRow(item)
    lv.setModel(model)


# get project_object info from listview
def read_check_list(lv, item_list, check_dict):
    item_list.clear()
    check_dict.clear()
    model = lv.model()
    if model is None:
        return
    for index in range(model.rowCount()):
        item = model.item(index)
        text = item.text()
        item_list.append(text)
        if item.checkState() == Qt.Checked:
            check_dict[text] = 1


def slot_qlv_check_list(ui, qlv):
    sl = qlv.selectedIndexes()
    if len(sl) < 2:
        return
    model = qlv.model()
    check = Qt.Checked
    if model.item(sl[0].row()).checkState() == Qt.Checked:
        check = Qt.Unchecked
    for idx in sl:
        item = model.item(idx.row())
        item.setCheckState(check)


def slot_project_list_filter(ui):
    filter_str = ui._qle_proj_filter.text()
    ui.fill_proj_list(filter_str)

def fill_dict(d, l_item):
    for item in l_item:
        if item in d:
            continue
        d[item] = 1


# get all avaliable filenames in output dir
def get_all_filenames(p_obj):
    l_case = p_obj._case
    l_ver = p_obj._ver
    dir_out = p_obj._dirOutput
    d_name = {}
    for case in l_case:
        for ver in l_ver:
            cur_dir = os.path.join(dir_out, case, ver)
            sub = utils.get_sub_dir(cur_dir)
            stem = utils.get_stem_list(cur_dir)
            fill_dict(d_name, sub)
            fill_dict(d_name, stem)
    return list(d_name.keys())


def get_all_ver_names(p_obj):
    l_case = p_obj._case
    dir_out = p_obj._dirOutput
    d_name = {}
    for case in l_case:
        cur_dir = os.path.join(dir_out, case)
        sub = utils.get_sub_dir(cur_dir)
        fill_dict(d_name, sub)
    return list(d_name.keys())

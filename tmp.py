# -*- coding: utf-8 -*-
import pdb
import re
import os
import sys
import cv2
import math
import openpyxl
import threading
import numpy as np
from PIL import Image
from operator import itemgetter
import docx
import pathlib
import codecs
from PyQt5.QtWidgets import QProgressDialog, QApplication, QWidget
from os.path import join, isdir

def read_config_list(config_str, pattern):
    lc = len(config_str)
    lp = len(pattern)
    if lc < lp:
        return None
    if config_str[0:lp] != pattern:
        return None
    return config_str[lp+1:].split(" ")


def read_compare_config(file_config):
    if not os.path.exists(file_config):
        return None
    case_list = []
    ver_list = []
    alg_list = []
    content = None
    with open(file_config, encoding='utf-8') as f:
        content = f.readlines()
    print(content)
    str_list = [l.strip() for l in content]
    for line in str_list:
        if line[0:3] == "cas":
            case_list = read_config_list(line, "cas")
        elif line[0:3] == "ver":
            ver_list = read_config_list(line, "ver")
        elif line[0:3] == "alg":
            alg_list = read_config_list(line, "alg")
    return case_list, ver_list, alg_list


def regex_cid(in_str):
    m = re.search('.*([\d]{8}[\d\+/\-]*).*', in_str)
    if m:
        return m.group(1)
    return ""


# sort and itemgetter
# t_list = [(1, 0, 2)]
# t_list.append((1, 3, 4))
# t_list.append((2, 3, 6))
# t_list.append((3, 1, 6))
# t_list.append((4, 3, 8))
# t_list.append((1, 4, 9))
# t_list.append((2, 1, 10))
# t_list.sort(key=itemgetter(0,1))
# print(t_list[0:2])


#image
#img_pil = Image.open("c:/data/xls/check/img_2.png")


def get_line_length(l):
    w = abs(l[0][2] - l[0][0]) 
    h = abs(l[0][1] - l[0][3])
    return w * w + h * h


h_pattern = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
v_pattern = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 40))


def find_lines(img_cv, pattern, color):
    dot_pattern = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5))
    gray = cv2.cvtColor(img_cv,cv2.COLOR_BGR2GRAY)
    #gray = cv2.dilate(gray, pattern)
    dst = cv2.Canny(gray, 50, 200, 3);
    dst = cv2.dilate(dst, dot_pattern)
    cdst = cv2.cvtColor(dst, cv2.COLOR_GRAY2BGR)
    lines = cv2.HoughLinesP(dst, 1, 3.1415926 / 720, 100, 0, 0)
    if lines is None or len(lines) < 1:
        print("no line found")
        return cdst
    for l in lines:
        cv2.line(cdst, (l[0][0], l[0][1]), (l[0][2], l[0][3]), color, 3, 2)
    return cdst
    


def find_max_line(img_name):
    img_cv = cv2.imread(img_name)
    gray = cv2.cvtColor(img_cv,cv2.COLOR_BGR2GRAY)
    kernel_size = 5
    #blur_gray = cv2.GaussianBlur(gray,(kernel_size, kernel_size),0)
    h_pattern = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
    v_pattern = cv2.getStructuringElement(cv2.MORPH_RECT, (5, 3))
    #v_pattern = np.ones((5,5),np.uint8)
    gray = cv2.dilate(gray, h_pattern)
    dst = cv2.Canny(gray, 50, 200, 3);
    dst = cv2.dilate(dst, v_pattern)
    #dst = gray
    cdst = cv2.cvtColor(dst, cv2.COLOR_GRAY2BGR)
    lines = cv2.HoughLinesP(dst, 1, 3.1415926 / 720, 100, 0, 0)
    if lines is None or len(lines) < 1:
        print("no line found")
        return None
    max_l = lines[0]
    max_len = 0
    for l in lines:
        cv2.line(cdst, (l[0][0], l[0][1]), (l[0][2], l[0][3]), (255, 0, 0), 3, 2)
        cur_len = get_line_length(l)
        if cur_len > max_len:
            max_l = l
            max_len = cur_len
    #cv2.line(cdst, (max_l[0][0], max_l[0][1]), (max_l[0][2], max_l[0][3]), (255, 255, 0), 3, 2)
    #Image.fromarray(cdst).save("d:/tmp/img_line.png")
    return max_l


def rotate_horizontal(img, l):
    h, w = img.shape[:2]
    angle = math.atan((l[0][3] - l[0][1]) / (l[0][2] - l[0][0]))
    angle_d = angle / math.pi * 180
    mat = cv2.getRotationMatrix2D((h/2, w/2), angle_d, 1)
    res = cv2.warpAffine(img, mat, (w, h))
    # 转化角度为弧度
    theta = angle
    # 计算高宽比
    hw_ratio = float(h) / float(w)
    # 计算裁剪边长系数的分子项
    tan_theta = np.tan(theta)
    numerator = np.cos(theta) + np.sin(theta) * np.tan(theta)
    # 计算分母中和高宽比相关的项
    r = hw_ratio if h > w else 1 / hw_ratio
    # 计算分母项
    denominator = r * tan_theta + 1
    # 最终的边长系数
    crop_mult = numerator / denominator
    # 得到裁剪区域
    w_crop = int(crop_mult * w)
    h_crop = int(crop_mult * h)
    x0 = int((w - w_crop) / 2)
    y0 = int((h - h_crop) / 2)
    return res[y0 : y0 + h_crop, x0 : x0 + w_crop]


def denoise_image(img):
    img_bw = 255*(cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) > 5).astype('uint8')

    se1 = cv2.getStructuringElement(cv2.MORPH_RECT, (5,5))
    se2 = cv2.getStructuringElement(cv2.MORPH_RECT, (2,2))
    mask = cv2.morphologyEx(img_bw, cv2.MORPH_CLOSE, se1)
    mask = cv2.morphologyEx(mask, cv2.MORPH_OPEN, se2)
    mask = np.dstack([mask, mask, mask]) / 255
    out = img * mask
    return out


#img = cv2.imread(img_name)

#res = denoise_image(img)
#res = cv2.medianBlur(img, 3)
def denoise_by_components(in_img, min_size):
    img = in_img
    grey = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    bw = cv2.adaptiveThreshold(cv2.bitwise_not(grey), 255,
                               cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 25, -2)
    nb_components, output, stats, centroids = cv2.connectedComponentsWithStats(bw, connectivity=4)
    sizes = stats[1:, -1]
    nb_components = nb_components - 1
    #for every component in the image, you keep it only if it's above min_size
    for i in range(0, nb_components):
        if sizes[i] <= min_size:
            img[output == i + 1] = 255
    return img

#img = denoise_by_components(img, 30)
#cv2.imwrite("c:/tmp/denoise.png", img)

def copy_sheet(path_from, path_to, sheet_name):
    wb_from = openpyxl.load_workbook(path_from) 
    sheet_list = wb_from.sheetnames
    if sheet_name not in sheet_list:
        print("Error! no sheet {} in file {}".format(sheet_name, excel_from))
        return
    ws_from = wb_from[sheet_name]

    wb = openpyxl.load_workbook(path_to)
    try:
        ws = wb[sheet_name]
        wb.remove(ws)
    except KeyError:
        pass
    sheet = wb.create_sheet(title=sheet_name,index=13)

    for rid, r in enumerate(ws_from.iter_rows(values_only=True)):
        for cid, cell_v in enumerate(r):
            sheet.cell(row=rid+1, column=cid+1).data_type = "s"
            sheet.cell(row=rid+1, column=cid+1).value = cell_v
    wb.save(path_to)
    print("{} copyed from {} to {}".format(sheet_name, path_from, path_to))


# img_name = "d:/tmp/error.png"
# h_pattern = cv2.getStructuringElement(cv2.MORPH_RECT, (40, 1))
# v_pattern = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 10))

# img = cv2.imread(img_name)
# res_h = find_lines(img, h_pattern, (255, 0, 0))
# cv2.imwrite("d:/tmp/line_h.png", res_h)
# res_v = find_lines(res_h, v_pattern, (0, 255, 0))
# cv2.imwrite("d:/tmp/line_v.png", res_v)


#copy_sheet("d:/tmp/广州第一分公司202001.xlsx", "d:/tmp/tmp_out.xlsx", "TB 202001")
# t_in = openpyxl.load_workbook("d:/tmp/tmp.xlsx")
# t_out = openpyxl.Workbook()
# ws_in = t_in.active
# txt = ws_in.cell(row=1, column = 1).value
# ws_out = t_out.active
# ws_out.cell(row=1, column = 1).data_type = "s"
# ws_out.cell(row=1, column = 1).value = txt
# t_out.save("d:/tmp/tmp_out.xlsx")

#thread test
# app = QApplication(sys.argv)
# app.setStyleSheet("QMessageBox { messagebox-text-interaction-flags: 5; }")
# w = QWidget()
# w.show()
# pg = QProgressDialog("Creating Screenshot...", "Stop", 0, 2, w)
# for i in range(0, 3):
#     #sp
#     pg.setValue(i)

# sys.exit(app.exec_())

def list_subdir_with_depth(root, depth=3):
    res_up = []
    res = [""]
    for d in range(0, depth):
        res_up = res.copy() # 1/2
        res = []
        for r in res_up:
            tr = join(root, r) # tmp root
            if len(r) < 1:
                res = res + [f for f in os.listdir(tr) if isdir(join(tr, f))]
            else:
                res = res + ["{}/{}".format(r, f) for f in os.listdir(tr) if isdir(join(tr, f))]
    return res
            

#print(list_subdir_with_depth("d:/tmp/tb"))

print(__file__)
print(os.path.join(os.path.dirname(__file__), '..'))
print(os.path.dirname(os.path.realpath(__file__)))
print(os.path.abspath(os.path.dirname(__file__)))

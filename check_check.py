# -*- coding: utf-8 -*-
## @file check_check.py
## @brief create diff from given check list file
## @author jiayanming
import pdb
import os.path
import re
import math
from openpyxl import *
#from openpyxl import styles
from bisect import bisect_left

dir_input = "c:/data/xls/1910/input/"
dir_output = "c:/data/xls/1910/output/"

my_color = styles.colors.Color(rgb="ffff00")
around_color = styles.fills.PatternFill(patternType='solid', fgColor=my_color)

# global variables
ver_map = {}
ap_map = {}
table_ver_input = load_workbook(os.path.join(dir_input, "ver.xlsx"))
table_ap_input = load_workbook(os.path.join(dir_input, "ap.xlsx"))

# global variable
ver_list = []
ap_input_list = []
ap_list = []
am_err_list = []
am_round_list = []
no_id_list = []
invalid_list = []
rm_ver_list = [] # ver list after filter equal check
sup_list = [] # suplier list without equal checks
sup_map = {} 

test_ver_list = [] # remain ver item after all filters, need test
test_ap_list = [] # remain ap item after all filters, need test

class CheckItem:
    def __init__(self, cid, amount, supplier, rid):
        self.check_id = cid
        self.amount = amount
        self.sup = supplier
        self.rid = rid
        self.map_id = -1
        self.dup_list = []
    def __lt__(self, other):
        if isinstance(other, self.__class__):
            return self.sup < other.sup
        return self.sup < str(other)
    
    check_id = 0 #use int for fast compare
    amount = 0 
    sup = "" # supplier
    rid = -1
    dup_list = []
    map_id = -1


# supplier item
# record total amount from both ap and ver
# @note already verified item excluded
class SupItem:
    def __init__(self, sup):
        self.sup = sup
        self.ver_am = 0
        self.ap_am = 0
        self.ver_count = 0
        self.ap_count = 0

def binary_search(a, x, lo=0, hi=None):
    hi = hi if hi is not None else len(a)
    pos = bisect_left(a, x, lo, hi)
    return (pos if pos != hi and a[pos].check_id == x else -1)


def merge_str(long_str, short_str):
    len_s = len(long_str) - len(short_str)
    return long_str[0:len_s] + short_str


## find previous number from given start position
def find_prev_number(input_str, pos):
    idx = pos - 1
    while idx >= 0 and input_str[idx].isdigit():
        idx -= 1
    return input_str[idx+1:pos]


## find next number from given start position
def find_next_number(input_str, pos):
    max_pos = len(input_str)
    idx = pos + 1
    while idx < max_pos and input_str[idx].isdigit():
        idx += 1
    return input_str[pos+1:idx]

## parse id list from compound str
## @return
## 0 normal finish
## 1 prefix invalid
## 2 prefix valid & no valid number following
## 3 ambiguous parse
def parse_check_id(ap_item, out_list=[]):
    input_id = ap_item.check_id
    out_list.clear()
    if len(input_id) <= 8:
        return 1
    prefix = input_id[0:8]
    #lenth judge for short ids
    if not prefix.isdigit():
        return 1
    if prefix in ver_map:
        ap_item.map_id = ver_map[prefix]
        ap_item.sup = ver_list[ap_item.map_id].sup
    
    next_number = find_next_number(input_id, 8)
    len_next = len(next_number)
    if len_next == 0 or len_next >= 8:
        out_list.append(prefix)
        return 2 # cannot find next number
    prev_number = prefix[8 - len_next:]
    if input_id[8] == "-":
        start = min(int(prev_number), int(next_number))
        end = max(int(prev_number), int(next_number))
        for i in range(start, end + 1):
            out_list.append(merge_str(prefix, str(i)))
    else:
        out_list.append(merge_str(prefix, prev_number))
        out_list.append(merge_str(prefix, next_number))
    # add other check
    split_pos = 8
    while True:
        split_pos = split_pos + len_next + 1
        if split_pos >= len(input_id):
            break
        next_number = find_next_number(input_id, split_pos)
        len_next = len(next_number)
        out_list.append(merge_str(prefix, next_number))
    return 0


def amount_equal(ver_list, ap_item, cid_list):
    ap_rid = ap_item.rid
    ap_am = ap_item.amount
    ver_ids = []
    ver_am = 0
    sup = ["JYM"]
    for i in cid_list:
        if i not in ver_map:
            continue  # skip item
        tmp_find = ver_map[i]
        ver_ids.append(tmp_find)
        ver_item = ver_list[tmp_find]
        if ver_item.map_id != -1:
            print("already mark v:{}, a:{}".
                  format(ver_item.rid, ap_rid))
        ver_am += ver_item.amount
        if sup[0] != "JYM" and sup[0] != ver_item.sup:
            print("ERROR! sup in ap table:{}".format(ap_rid))
        sup[0] = ver_item.sup
    ap_item.sup = sup[0]
    if not math.isclose(ap_am, ver_am, abs_tol=0.995):
        return False
    # record map information
    if math.isclose(ap_am, ver_am, abs_tol=1e-5):
        ap_item.map_id = 0
    else:
        ap_item.map_id = -2
    if len(ver_ids) == 1:
        ap_item.map_id = ver_list[ver_ids[0]].rid
    for i in ver_ids:
        if ver_list[i].map_id != -1:
            print("already set:v-{}, a-{}".format(i, ap_rid))
        ver_list[i].map_id = ap_rid
    return True


## @todo return map information
## return ver_ids specify verified item positions in ver_list
def id_group_equal(ver_list, ap_item):
    org_id_list = []
    id_type = parse_check_id(ap_item, org_id_list)
    id_list = list(set(org_id_list)) # remove duplicated
    if id_type == 0:
        return amount_equal(ver_list, ap_item, id_list)
    if id_type == 1:
        return False
    if id_type == 2:
        return amount_equal(ver_list, ap_item, id_list[0:1])
    if id_type == 3:
        res_p = amount_equal(ver_list, ap_item, id_list[-2:-1])
        if res_p is True:  # if prefix equal
            return True
        res_l2 = amount_equal(ver_list, ap_item, id_list[-1:])
        if res_l2 is True:  # if last 2 equal
            return True
        res_n2 = amount_equal(ver_list, ap_item, id_list[:-2])
        if res_n2 is True:  # if first n-2 equal
            return True
    return False

# ##################### FILE LOAD ####################    
# column
#          check  amount  sup
# verify :   B      E      H
# input  :   F      M      query in verify
def load_verify_item(ver_table, ver_list):
    ver_map.clear()
    ws = ver_table.active
    rid = 2
    for r in ws.iter_rows(min_row=2, max_col=8, values_only=True):
        cid = r[1]
        cur_am = float(r[4])
        cur_sup = r[7]
        ver_map[cid] = len(ver_list)
        ver_list.append(CheckItem(cid, cur_am, cur_sup, rid))
        rid += 1


# fetch check id from string
def regex_cid(in_str):
    m = re.search('.*([\d]{8}[\d\+/\-]*).*', in_str)
    if m:
        return m.group(1)
    return ""


# load ap item and remove duplicated
def load_ap_input(ap_table, ap_input_list):
    ws = ap_table.active
    rid = 1
    #pdb.set_trace()
    for r in ws.iter_rows(min_row=2, max_col=11, values_only=True):
        rid += 1
        cid = ""
        # find in column 15 text
        text_str = str(r[5]).replace("~", "-")
        cur_ref = str(r[4]).replace("~", "-")
        am = float(r[10])

        if text_str is None:
            print(r)
        if len(text_str) >= 8:
            regex_res = regex_cid(text_str)
            if len(regex_res) != 0:
                cid = regex_res
        # find in reference
        if len(cid) < 8:
            cid = cur_ref.replace(" ", "")
            if len(cid) == 10 and (cid[8] == "-"):
                cid = cid[0:8]
        if cid not in ap_map:
            ap_map[cid] = len(ap_input_list)
            ap_input_list.append(CheckItem(cid, am, "", rid))
        else:
            old_pos = ap_map[cid]
            old_item = ap_input_list[old_pos]
            old_item.amount += am
            old_item.dup_list.append(rid)
# end##################### FILE LOAD ####################    

## load ap check items
## @return:
## ap_list varified list
##         map_id: 0  -- correspond to multiple ver item
##                 -2 -- net amount is 0
##                 -1 -- n
## am_err_list check item found in ver_list but amount not equal
## no_id_list  check_item cannot found in ver_list
## invalid_list check_item with invalid check_id
## @note need ver_list filled first
def filter_ap_item():
    for cur_item in ap_input_list:
        am = cur_item.amount
        # net
        if math.isclose(am, 0, abs_tol=1e-5):
            cur_item.map_id = -2
            ap_list.append(cur_item)
            continue
        rid = cur_item.rid
        cid = cur_item.check_id
        if len(cid) != 8 or not (cid.isdigit()):
            if id_group_equal(ver_list, cur_item):
                ap_list.append(cur_item)
                continue
            else:
                invalid_list.append(cur_item)
                continue
        if cid not in ver_map:
            no_id_list.append(cur_item)
            continue
        ver_item = ver_list[ver_map[cid]]
        cur_item.sup = ver_item.sup
        if math.isclose(ver_item.amount, am, abs_tol=0.995):
            ver_item.map_id = rid
            cur_item.map_id = ver_item.rid
            if math.isclose(ver_item.amount, am, abs_tol=1e-5):
                ap_list.append(cur_item)
            else:
                am_round_list.append(cur_item)
            continue
        am_err_list.append(cur_item)


## debugging methods
def print_ver_number(ver_list, stage):
    verified_number = 0
    for v in ver_list:
        if v.map_id != -1:
            verified_number += 1
    print("verified number at {} : {}".format(stage, verified_number))


def write_item_to_file(filename, out_list):
    wb = Workbook()
    ws = wb.active
    ws.append(["check_id", "amount", "row_no", "map_id", "supplier"])
    for vi in out_list:
        ws.append([vi.check_id, vi.amount, vi.rid, vi.map_id, vi.sup])
    wb.save(filename)


################ supplier relevant
def get_remain_ver_item(ver_list, rm_ver_list):
    for vi in ver_list:
        if vi.map_id != -1:
            continue
        rm_ver_list.append(vi)


# get amount sum for each supplier,
# return sum list and map from sup_NAME to list_IDX
def collect_supplier_sum(item_list, list_sum, sup_map):
    i = 0
    total = len(item_list)
    while i < total:
        cur_sup = item_list[i].sup
        sum_am = 0
        cur_start = i
        # collect amount of same cid
        for j in range(cur_start, total):
            if cur_sup != item_list[j].sup:
                break
            sum_am += item_list[j].amount
            i += 1
        # mark verified
        sup_map[cur_sup] = len(list_sum)
        list_sum.append(sum_am)


# write supplier with different number
def write_item_to_sup(filename, sup_list):
    wb = Workbook()
    ws = wb.active
    ws.append(["sup_name", "ver_amount", "ap_amount", "ver_count", "ap_count"])
    for vi in sup_list:
        if math.isclose(vi.ap_am, vi.ver_am, abs_tol=1e-5):
            continue
        ws.append([vi.sup, vi.ver_am, vi.ap_am, vi.ver_count, vi.ap_count])
    wb.save(filename)


########### update output################
# 0: not equal
# 1: equal
# 2: round equal
def sup_status(sup_name):
    si = sup_list[sup_map[sup_name]]
    if math.isclose(si.ap_am, si.ver_am, abs_tol=0.995):
        if math.isclose(si.ap_am, si.ver_am, abs_tol=1e-5):
            return 1
        return 2
    return 0


# amount same, mark id
def mark_same_id(cur_item, ws, col, value, color=False):
    if cur_item.amount == 1:
        return
    for ri in cur_item.dup_list:
        ws.cell(row=ri, column=col).value = value
        if color:
            ws.cell(row=ri, column=col).fill = around_color


# write res to ver table
def update_ver_table():
    ws = table_ver_input.active
    ws.insert_cols(1)
    ws.cell(row=1, column=1).value = "res"
    for item in ver_list:
        ws.cell(row=item.rid, column=1).value = item.map_id
    for item in rm_ver_list:
        sup_st = sup_status(item.sup)
        if sup_st == 0:
            test_ver_list.append(item)
            continue
        ws.cell(row=item.rid, column=1).value = "VENDOR"
        if sup_st == 2:
            ws.cell(row=item.rid, column=1).fill = around_color


# write results to ap table 
def update_ap_table():
    ws = table_ap_input.active
    ws.insert_cols(1)
    ws.insert_cols(2)
    ws.cell(row=1, column=1).value = "res"
    for item in ap_list:
        res = item.map_id
        ws.cell(row=item.rid, column=2).value = item.sup
        mark_same_id(item, ws, 2, item.sup)
        if res == -1:
            continue
        if res == -2:
            ws.cell(row=item.rid, column=1).value = "NET"
            mark_same_id(item, ws, 1, "NET")
            continue
        ws.cell(row=item.rid, column=1).value = res
        mark_same_id(item, ws, 1, res)
    # around same
    for item in am_round_list:
        ws.cell(row=item.rid, column=1).value = item.map_id
        mark_same_id(item, ws, 1, item.map_id)
        ws.cell(row=item.rid, column=1).fill = around_color
        ws.cell(row=item.rid, column=2).value = item.sup
        mark_same_id(item, ws, 2, item.sup)
    # amount incorrect
    for item in am_err_list:
        ws.cell(row=item.rid, column=2).value = item.sup
        mark_same_id(item, ws, 2, item.sup)
        sup_st = sup_status(item.sup)
        if sup_st == 0:
            test_ap_list.append(item)
            continue
        ws.cell(row=item.rid, column=1).value = "VENDOR"
        mark_same_id(item, ws, 1, "VENDOR")
        if sup_st == 2:
            ws.cell(row=item.rid, column=1).fill = around_color
            mark_same_id(item, ws, 1, "VENDOR", color=True)


def calculate_sup_sum():
    for item in rm_ver_list:
        sup = item.sup
        if sup not in sup_map:
            next_number = len(sup_list)
            sup_map[sup] = next_number
            sup_list.append(SupItem(sup))
            sup_list[next_number].ver_am = item.amount
            sup_list[next_number].ver_count += 1
        else:
            sup_list[sup_map[sup]].ver_am += item.amount
            sup_list[sup_map[sup]].ver_count += 1

    for item in am_err_list:
        sup = item.sup
        if sup not in sup_map:
            next_number = len(sup_list)
            sup_map[sup] = next_number
            sup_list.append(SupItem(sup))
            sup_list[next_number].ap_am = item.amount
            sup_list[next_number].ap_count += 1
        else:
            sup_list[sup_map[sup]].ap_am += item.amount
            sup_list[sup_map[sup]].ap_count += 1
    

##############debug##########################
def get_am_sum(l):
    sum = 0
    for item in l:
        sum += item.amount
    return sum

############## start process ########################

# read input and make basic compare
load_verify_item(table_ver_input, ver_list)
#print("ver sum: {}".format(get_am_sum(ver_list)))
load_ap_input(table_ap_input, ap_input_list)
#print("ap sum: {}".format(get_am_sum(ap_input_list)))
filter_ap_item()
# get remain items and compare by supplier
get_remain_ver_item(ver_list, rm_ver_list)
calculate_sup_sum()
############ debug ################
#write_item_to_file(os.path.join(dir_output, "ap_dup.xlsx"), ap_input_list)
# write results
update_ver_table()
update_ap_table()
table_ap_input.save(os.path.join(dir_output, "ap_output.xlsx"))
table_ver_input.save(os.path.join(dir_output, "verify_output.xlsx"))
write_item_to_file(os.path.join(dir_output, "invalid_id.xlsx"), invalid_list)
write_item_to_file(os.path.join(dir_output, "remain_ap.xlsx"), test_ap_list)
write_item_to_file(os.path.join(dir_output, "remain_ver.xlsx"), test_ver_list)
write_item_to_sup(os.path.join(dir_output, "remain_sup.xlsx"), sup_list)

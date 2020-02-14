# -*- coding: utf-8 -*-
## @author jiayanming
import pdb
import shutil
import os.path
import sys
import re
import math
from openpyxl import *

## constants
yello = styles.colors.Color(rgb="ffff00")
red = styles.colors.Color(rgb="ff0000")
white = styles.colors.Color(rgb="ffffff")
yello_color = styles.fills.PatternFill(patternType='solid', fgColor=yello)
red_color = styles.fills.PatternFill(patternType='solid', fgColor=red)
white_color = styles.fills.PatternFill(patternType='solid', fgColor=white)
bias = 2

idx_avalible = 5 + bias
idx_hz = 11 + bias
idx_health = 18 + bias
idx_dep = 4 + bias

## unhealth
idx_name = 0
idx_sick = 6 + bias
idx_rel_sick = 7 + bias
idx_region_return = 10 + bias
idx_region_pass = 9 + bias

idx_travel = 21 + bias

# idx_name_list = [
#     "提交人", "部门", "是否填写", "本人目前有无发烧、咳嗽等症状", "直系亲属目前有无发烧、咳嗽等症状", "最近14天，有无与湖北、温州乐清、台州（黄岩、温岭）以及河南（信阳、南阳）、安徽（安庆）疫情严重地区人员的接触史？", "是否将从湖北、温州乐清、台州（黄岩、温岭）以及河南（信阳、南阳）、安徽（安庆）返回？", "是否已回到杭州" "是否已经取得“健康码”"
# ]
# def update_idx(r):
#     idx_name = r.find(idx_name_list[0])
#     idx_dep = r.find(idx_name_list[1])
#     idx_avalible = find(idx_name_list[2])
#     idx_sick = find(idx_name_list[3])
#     idx_rel_sick = find(idx_name_list[4])
#     idx_region_pass = find(idx_name_list[5])
#     idx_region_return = find(idx_name_list[6])
#     idx_hz = find(idx_name_list[7])
#     idx_health = find(idx_name_list[8])

dep_name_list = [
    "董事会办公室",
    "财务部","国际事业部",
"知识产权法务部", 
"科技发展部", 
"行政部", 
"人力资源部", 
"信息技术部", 
"品质部", 
"采购物流部", 
"生产部", 
"杭州研究院", 
"研发管理部", 
"基础软件研发部", 
"基础硬件研发部", 
"测试部", 
"3D打印研发部", 
"3D数字化研发部", 
"齿科数字化研发部", 
"销售管理部", 
"市场部", 
"齿科数字化事业部", 
"3D数字化渠道部", 
"3D鞋数字化产品部", 
"审计部", 
"数字系统"]

# class
class colleage:
    def __init__(self, name):
        self.name = name
        self.dep = ""
        self.nodata = False
        self.health = ""
        self.sick = False
        self.rel_sick = False
        self.region_return = False
        self.region_pass = False
        self.travel = False
    def is_danger(self):
        return self.sick or self.rel_sick or self.region_return or self.region_pass or self.health != my_decode("绿色")


class depart:
    def __init__(self, name):
        self.dep_name = name
        self.total = 0
        self.hz = 0
        self.hz_r = 0
        self.hz_y = 0
        self.hz_g = 0
        self.hz_n = 0
        self.nhz = 0
        self.nhz_r = 0
        self.nhz_y = 0
        self.nhz_g = 0
        self.nhz_n = 0
        self.nodata = 0

# global variables
dep_map = {}
dep_list = []

list_normal = []
list_nodata = []
list_danger = []

list_cand_name = []

def my_decode(org):
    #return bytes(org, 'utf-8-sig').decode("gbk","ignore")
    #return org.encode('gbk').decode("utf-8", 'ignore')
    return org

def filter_cand_table(filename):
    for p in list_nodata:
        p.name

# process one file
def run_st(filename):
    dep_map.clear()
    dep_list.clear()
    list_nodata.clear()
    list_danger.clear()
    list_normal.clear()
    #### load file
    table = load_workbook(filename)
    ws = table.active
    #### read lines
    rows =  ws.iter_rows(min_row=2, max_col=23 + bias, values_only=True)
    rid = 0
    for r in rows:
        rid += 1
        # 部门, 有多个部门时，取最前
        tmp_dep = my_decode(r[idx_dep])
        dep = ""
        dep_pos = 1000
        for dn in dep_name_list:
            cur_pos = tmp_dep.find(dn)
            if cur_pos == -1 or cur_pos > dep_pos:
                continue
            dep = dn
            dep_pos = cur_pos
        if dep == "":
            print("warning! cannot find dep for line{}".format(rid))
        # 其它信息
        nodata = my_decode(r[idx_avalible]) == my_decode("否")
        hz_str = my_decode(r[idx_hz])
        hz = hz_str  == my_decode("已返回")
        health = my_decode(r[idx_health])
        travel = my_decode(r[idx_travel]) == my_decode("5-以上均不符合")
        name = my_decode(r[idx_name]).split('/')[0]
        cur_person = colleage(name)
        cur_person.dep = dep
        cur_person.nodata = nodata
        cur_person.health = health
        cur_person.sick = my_decode(r[idx_sick]) == my_decode("有")
        cur_person.rel_sick = my_decode(r[idx_rel_sick]) == my_decode("有")
        cur_person.region_return = my_decode(r[idx_region_return]) == my_decode("是")
        cur_person.region_pass = my_decode(r[idx_region_pass]) == my_decode("是")
        cur_person.travel = travel
        if nodata:
            list_nodata.append(cur_person)
        elif cur_person.is_danger():
            list_danger.append(cur_person)
        else:
            list_normal.append(cur_person)
        ## info
        cur_dep = None
        if dep in dep_map:
            cur_dep = dep_list[dep_map[dep]]
        else:
            tmp_id = len(dep_list)
            dep_map[dep] = tmp_id
            dep_list.append(depart(dep))
            cur_dep = dep_list[tmp_id]
        cur_dep.total += 1
        if nodata:
            cur_dep.nodata += 1
            continue
        if hz:
            cur_dep.hz += 1
            if health == my_decode("绿色"):
                cur_dep.hz_g += 1
            elif health == my_decode("黄色"):
                cur_dep.hz_y += 1
            elif health == my_decode("红色"):
                cur_dep.hz_r += 1
            else:
                cur_dep.hz_n += 1
        else:
            cur_dep.nhz += 1
            if health == my_decode("绿色"):
                cur_dep.nhz_g += 1
            elif health == my_decode("黄色"):
                cur_dep.nhz_y += 1
            elif health == my_decode("红色"):
                cur_dep.nhz_r += 1
            else:
                cur_dep.nhz_n += 1


# write st results
def write_table(out_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["部门", "总数A", "未填写", "在杭B", "在杭绿码C", "在杭黄码D", "在杭红码E", "未申请码X", "未回杭F", "未回杭绿码G", "回杭黄码H", "未回杭红码I", "未回未申请码Y"])
    for dep in dep_list:
        ws.append([
        dep.dep_name,
        dep.total,
        dep.nodata,
        dep.hz,
        dep.hz_g,
        dep.hz_y,
        dep.hz_r,
        dep.hz_n,
        dep.nhz,
        dep.nhz_g,
        dep.nhz_y,
        dep.nhz_r,
        dep.nhz_n
        ])
    wb.save(out_file)


def write_nodata_table(filename):
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "部门"])
    for p in list_nodata:
        ws.append([p.name, p.dep])
    wb.save(filename)

def write_danger_table(filename):
    wb = Workbook()
    ws = wb.active
    ws.append(["姓名", "部门", "健康码", "生病", "疫区接触"])
    rid = 1
    for p in list_danger:
        rid += 1
        sick = "否"
        if p.sick or p.rel_sick:
            sick = "是"
        contact = "否"
        if p.region_return or p.region_pass:
            contact = "是"
        ws.append([p.name, p.dep,
                   p.health,
                   sick, contact
        ])
        if p.health != "绿色":
            ws.cell(row=rid, column=3).fill = yello_color
        if sick == "是":
            ws.cell(row=rid, column=4).fill = yello_color
        if contact == "是":
            ws.cell(row=rid, column=5).fill = yello_color

    wb.save(filename)    


def fill_line(p, ws, rid):
    if p.nodata:
        ws.cell(row = rid, column = 3).value = "未填写"
    if p.health != "绿色":
        ws.cell(row = rid, column = 4).value = p.health
    if p.sick or p.rel_sick:
        ws.cell(row = rid, column = 5).value = "生病"
    if p.region_return or p.region_pass:
        ws.cell(row = rid, column = 6).value = "疫区接触"
    if p.travel:
        ws.cell(row = rid, column = 7).value = "不满足入园条件"



def highlight_cand(filename, out_tmp_name):
    name_map = {}
    #### load file
    table = load_workbook(filename)
    ws = table.active
    ## insert col
    # ws.insert_cols(2)
    # ws.insert_cols(3)
    # ws.insert_cols(4)
    # ws.insert_cols(5)
    #### read lines
    rows =  ws.iter_rows(min_row=4, max_col=20, values_only=True)
    rid = 4
    ws.cell(row = 2, column = 3).value = "未填写"
    ws.cell(row = 2, column = 4).value = "健康码"
    ws.cell(row = 2, column = 5).value = "生病"
    ws.cell(row = 2, column = 6).value = "疫区接触"
    ws.cell(row = 2, column = 7).value = "入园条件"

    for r in rows:
        name_map[r[1]] = rid
        ws.cell(row = rid, column = 2).fill = yello_color
        ws.cell(row = rid, column = 3).value = ""
        ws.cell(row = rid, column = 4).value = ""
        ws.cell(row = rid, column = 5).value = ""
        ws.cell(row = rid, column = 6).value = ""
        ws.cell(row = rid, column = 7).value = ""
        rid += 1

    for n in list_nodata:
        if n.name in name_map:
            rid = name_map[n.name]
            ws.cell(row = rid, column = 2).fill = red_color
            fill_line(n, ws, rid)
    for n in list_danger:
        if n.name in name_map:
            rid = name_map[n.name]
            ws.cell(row = rid, column = 2).fill = red_color
            fill_line(n, ws, rid)
    for n in list_normal:
        if n.name in name_map:
            rid = name_map[n.name]
            if n.travel:
                ws.cell(row = rid, column = 2).fill = red_color
            else:                
                ws.cell(row = rid, column = 2).fill = white_color
            fill_line(n, ws, rid)
    table.save(out_tmp_name)

def get_white_cand(w_dir, out_dir):
    dir_input = os.fsencode(w_dir)
    for file in os.listdir(dir_input):
        if not os.path.isfile(os.path.join(dir_input, file)):
            continue
        filename = os.fsdecode(file)
        stem, ext = os.path.splitext(filename)
        if ext != ".xlsx":
            continue
        full_filename = os.path.join(w_dir, filename)
        tmp_filename = os.path.join(w_dir, "tmp_cand.xlsx")
        tmp_out_filename = os.path.join(w_dir, "tmp_cand_out.xlsx")
        out_filename = os.path.join(out_dir, stem + "_out.xlsx")
        shutil.move(full_filename, tmp_filename)
        highlight_cand(tmp_filename, tmp_out_filename)
        shutil.move(tmp_filename, full_filename)
        if os.path.exists(out_filename):
            os.remove(out_filename)
        shutil.move(tmp_out_filename, out_filename)


# main: process all xlsx file in current dir
cwd = os.getcwd()
str_input = os.path.join(cwd, "input")
str_output = os.path.join(cwd, "output")

dir_input = os.fsencode(str_input)
white_dir = os.path.join(cwd, "names")

# read candidate list
#get_white_cand(white_dir)

for file in os.listdir(dir_input):
    if not os.path.isfile(os.path.join(dir_input, file)):
        continue
    filename = os.fsdecode(file)
    stem, ext = os.path.splitext(filename)
    if ext != ".xlsx":
        continue
    full_filename = os.path.join(str_input, filename)
    tmp_filename = os.path.join(str_input, "tmp.xlsx")
    shutil.move(full_filename, tmp_filename)
    run_st(tmp_filename)
    shutil.move(tmp_filename, full_filename)
    
    out_filename = os.path.join(str_output, stem + "_out.xlsx")
    nodata_filename = os.path.join(str_output, stem + "_no_data.xlsx")
    danger_filename = os.path.join(str_output, stem + "_danger.xlsx")
    if os.path.exists(out_filename):
        os.remove(out_filename)
    if os.path.exists(nodata_filename):
        os.remove(nodata_filename)
    if os.path.exists(danger_filename):
        os.remove(danger_filename)
    write_table(tmp_filename)
    shutil.move(tmp_filename, out_filename)
    write_nodata_table(tmp_filename)
    shutil.move(tmp_filename, nodata_filename)
    write_danger_table(tmp_filename)
    shutil.move(tmp_filename, danger_filename)

get_white_cand(white_dir, str_output)

